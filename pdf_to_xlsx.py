#!/usr/bin/env python3
"""
pdf_to_xlsx.py — Converte PDFs de ensaios GENVCE em XLSX formatado.

Uso:
    python pdf_to_xlsx.py <arquivo.pdf> [saida.xlsx] [--logo logo.png]

Exemplos:
    python pdf_to_xlsx.py ensaio.pdf
    python pdf_to_xlsx.py ensaio.pdf resultado.xlsx
    python pdf_to_xlsx.py ensaio.pdf resultado.xlsx --logo genvce.jpeg
"""

import sys
import re
import argparse
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    sys.exit("Erro: instale pdfplumber com:  pip install pdfplumber")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Erro: instale openpyxl com:  pip install openpyxl")


# ── Estilos ───────────────────────────────────────────────────────────────────

GREEN_HDR  = '388E3C'
GREEN_EVEN = 'F1F8E9'
WHITE      = 'FFFFFF'
GRAY_TEXT  = '444444'

def fill(color):
    return PatternFill('solid', fgColor=color)

def style_header_cell(cell, text, num_cols=1):
    cell.value = text
    cell.font = Font(bold=True, size=10, color='FFFFFF', name='Arial')
    cell.fill = fill(GREEN_HDR)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_data_cell(cell, value, zebra=False, align='center', bold=False, size=10):
    cell.value = value
    cell.font = Font(size=size, bold=bold, name='Arial', color=GRAY_TEXT if not bold else '000000')
    cell.fill = fill(GREEN_EVEN if zebra else WHITE)
    cell.alignment = Alignment(horizontal=align, vertical='center')

def style_stat_cell(cell, value, bold=False, align='left'):
    cell.value = value
    cell.font = Font(size=10, bold=bold, name='Arial')
    cell.alignment = Alignment(horizontal=align, vertical='center')

def style_meta_cell(cell, value, size=9, bold=False):
    cell.value = value
    cell.font = Font(size=size, bold=bold, name='Arial')
    cell.alignment = Alignment(horizontal='left', vertical='center')


# ── Extração do PDF ───────────────────────────────────────────────────────────

def try_numeric(val):
    """Tenta converter string para int ou float."""
    if val is None:
        return None
    s = str(val).strip()
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s.replace(',', '.'))
    except ValueError:
        pass
    return s

def parse_page(page):
    """
    Retorna dict com:
      - title:   título do documento (cabeçalho repetido em todas as páginas)
      - section: subtítulo/seção da página
      - headers: lista de cabeçalhos da tabela
      - rows:    lista de listas com dados
      - stats:   lista de (label, valor) para bloco de estatísticas
      - footnote: texto de rodapé
    """
    raw_tables = page.extract_tables()
    full_text  = page.extract_text() or ''
    lines      = [l.strip() for l in full_text.splitlines() if l.strip()]

    result = {
        'title':    lines[0] if lines else '',
        'section':  '',
        'headers':  [],
        'rows':     [],
        'stats':    [],
        'footnote': '',
    }

    if not raw_tables:
        # Página só de texto
        result['section'] = lines[1] if len(lines) > 1 else ''
        return result

    table = raw_tables[0]

    # Identifica seção e cabeçalho da tabela
    data_start = 0
    for i, row in enumerate(table):
        non_null = [c for c in row if c and str(c).strip()]
        # Linha de seção: só col 0 preenchida, sem dados numéricos
        if len(non_null) == 1 and i <= 1:
            text = non_null[0].strip()
            # Ignora blocos de texto longo (metadados do documento)
            if len(text.splitlines()) == 1 and len(text) < 80:
                result['section'] = text
            data_start = i + 1
            continue
        # Linha de cabeçalho real: primeira célula vazia, resto preenchido
        if not (row[0] or '').strip() and any(row[1:]):
            result['headers'] = ['VARIEDAD'] + [c.strip() for c in row[1:] if c]
            data_start = i + 1
            break

    # Separa linhas de dados das linhas de estatísticas
    stat_labels = {'media', 'desviación', 'coeficiente', 'diseño', 'variedad testigo'}
    for row in table[data_start:]:
        if not any(row):
            continue
        first = str(row[0] or '').strip()
        if not first:
            continue
        # Rodapé
        if first.startswith('*') and ':' in first:
            result['footnote'] = first
            continue
        # Estatística
        if any(s in first.lower() for s in stat_labels):
            val = try_numeric(row[1]) if len(row) > 1 else None
            result['stats'].append((first, val if val is not None else (row[1] or '')))
            continue
        # Linha de dados
        result['rows'].append([try_numeric(c) if i > 0 else (c or '') for i, c in enumerate(row)])

    # Se seção não encontrada na tabela, busca nas linhas de texto
    if not result['section']:
        for line in lines:
            # Linha que parece título de seção: não contém datas, não é o título principal, não é nome de variedad
            if (len(line) > 10 and len(line) < 100
                    and not re.search(r'\d{4}|fecha|ensayo en|instituto|iriaf|itacyl', line.lower())
                    and not line.isupper()):
                result['section'] = line
                break

    # Rodapé no texto da página se não encontrado na tabela
    if not result['footnote']:
        for line in reversed(lines):
            if line.startswith('*') and 'testigo' in line.lower():
                result['footnote'] = line
                break

    return result


# ── Escrita do XLSX ───────────────────────────────────────────────────────────

def write_sheet(ws, page_data, doc_meta, logo_path=None):
    """Escreve uma aba do XLSX com base nos dados extraídos da página."""

    # Linha 1: logo (se disponível) + título do documento no cabeçalho
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 6
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 6

    if logo_path and Path(logo_path).exists():
        img = XLImage(logo_path)
        img.width  = 120
        img.height = 23
        img.anchor = 'A1'
        ws.add_image(img)

    # Metadados (linhas 5-7)
    title_text = page_data['title'] or doc_meta.get('title', '')
    if page_data['section']:
        title_text = f"{doc_meta.get('base_title', title_text)} – {page_data['section']}"

    ws.row_dimensions[5].height = 18
    style_meta_cell(ws['A5'], title_text, size=13, bold=True)

    ws.row_dimensions[6].height = 13
    style_meta_cell(ws['A6'], doc_meta.get('instituto', ''), size=9)

    ws.row_dimensions[7].height = 13
    style_meta_cell(ws['A7'], doc_meta.get('dates', ''), size=9)

    # Subtítulo da seção (linha 8) se diferente do título
    if page_data['section']:
        ws.row_dimensions[8].height = 16
        style_meta_cell(ws['A8'], page_data['section'], size=11, bold=True)

    # Cabeçalho da tabela (linha 9)
    HDR_ROW = 9
    headers = page_data['headers']
    if not headers:
        return

    ws.row_dimensions[HDR_ROW].height = 30
    for c, h in enumerate(headers, 1):
        style_header_cell(ws.cell(HDR_ROW, c), h)

    # Dados
    for i, row in enumerate(page_data['rows']):
        r = HDR_ROW + 1 + i
        zebra = i % 2 != 0
        ws.row_dimensions[r].height = 15
        for c, val in enumerate(row):
            align = 'left' if c == 0 else 'center'
            style_data_cell(ws.cell(r, c + 1), val, zebra=zebra, align=align)

    last_data = HDR_ROW + len(page_data['rows'])

    # Estatísticas
    if page_data['stats']:
        r = last_data + 2
        for label, val in page_data['stats']:
            style_stat_cell(ws.cell(r, 1), label, bold=True)
            style_stat_cell(ws.cell(r, 2), val, align='center')
            r += 1

    # Rodapé
    footnote_row = last_data + len(page_data['stats']) + 3
    if page_data['footnote']:
        ws.cell(footnote_row, 1).value = page_data['footnote']
        ws.cell(footnote_row, 1).font = Font(size=9, name='Arial')

    # Larguras de coluna automáticas
    num_cols = len(headers)
    for c in range(1, num_cols + 1):
        col_letter = get_column_letter(c)
        lengths = [len(str(headers[c-1])) if headers[c-1] else 0]
        lengths += [len(str(row[c-1])) if c-1 < len(row) and row[c-1] else 0
                    for row in page_data['rows']]
        max_len = max(lengths) if lengths else 10
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)


def extract_doc_meta(first_page_text):
    """Extrai metadados gerais do documento a partir da primeira página."""
    lines = [l.strip() for l in first_page_text.splitlines() if l.strip()]
    meta = {
        'title':       lines[0] if lines else '',
        'base_title':  lines[0] if lines else '',
        'instituto':   '',
        'dates':       '',
    }
    date_parts = []
    for line in lines:
        if 'instituto' in line.lower() or 'iriaf' in line.lower() or 'itacyl' in line.lower():
            meta['instituto'] = line
        if re.search(r'fecha|siembra|cosecha', line.lower()):
            date_parts.append(line)
    if date_parts:
        meta['dates'] = '  |  '.join(date_parts)
    return meta


# ── Ponto de entrada ──────────────────────────────────────────────────────────

def convert(pdf_path, xlsx_path, logo_path=None):
    pdf_path  = Path(pdf_path)
    xlsx_path = Path(xlsx_path)

    if not pdf_path.exists():
        sys.exit(f"Erro: arquivo não encontrado: {pdf_path}")

    print(f"📄 Lendo: {pdf_path.name}")

    with pdfplumber.open(pdf_path) as pdf:
        pages_data = [parse_page(p) for p in pdf.pages]
        first_text = pdf.pages[0].extract_text() or ''

    doc_meta = extract_doc_meta(first_text)
    print(f"   Documento: {doc_meta['title']}")
    print(f"   Páginas encontradas: {len(pages_data)}")

    wb = Workbook()
    wb.remove(wb.active)  # remove aba default

    for i, pdata in enumerate(pages_data):
        # Nome da aba: seção ou "Página N"
        sheet_name = pdata['section'] or f'Página {i+1}'
        sheet_name = sheet_name[:31]  # limite do Excel
        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, pdata, doc_meta, logo_path=logo_path)
        cols = len(pdata['headers'])
        rows = len(pdata['rows'])
        print(f"   ✓ Aba '{sheet_name}': {cols} colunas, {rows} linhas de dados")

    wb.save(xlsx_path)
    print(f"\n✅ Salvo em: {xlsx_path}")


def main():
    parser = argparse.ArgumentParser(
        description='Converte PDFs de ensaios agrícolas (GENVCE) em XLSX formatado.'
    )
    parser.add_argument('pdf',  help='Caminho do arquivo PDF de entrada')
    parser.add_argument('xlsx', nargs='?', help='Caminho do arquivo XLSX de saída (opcional)')
    parser.add_argument('--logo', help='Caminho de uma imagem de logo para inserir nas abas (PNG ou JPEG)')
    args = parser.parse_args()

    pdf_path  = Path(args.pdf)
    xlsx_path = Path(args.xlsx) if args.xlsx else pdf_path.with_suffix('.xlsx')

    convert(pdf_path, xlsx_path, logo_path=args.logo)


if __name__ == '__main__':
    main()
